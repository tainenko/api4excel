#todo:replace the xlrd library with openpyxl
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import xlrd
import xlutils.copy
from Base.readConfig import ReadConfig
import time

class ReadExcel:
    def __init__(self,section,field,sheet):
        # 打开工作表，并定位到sheet
        self.data_address = ReadConfig().get_config(section,field)
        self.workbook=load_workbook(data_address)
        workbook = xlrd.open_workbook(data_address)
        self.table = workbook.sheets()[sheet]

    @property
    def get_sheetnames(self):
        return self.workbook.sheetnames


'''
    def get_rows(self):
        # 获取excel行数
        rows = self.table.nrows
        return rows

    def get_cell(self,row,col):
        # 获取单元格数据
        cell_data = self.table.cell(row,col).value
        return cell_data

    def get_col(self,col):
        # 获取整列数据
        col_data = self.table.col_value(col)
        return col_data


class WriteExcel:
    def __init__(self,section,field,sheet):
        # 打开工作表
        self.address = ReadConfig().get_config(section,field)
        self.workbook = xlrd.open_workbook(self.address)
        self.wf = xlutils.copy.copy(self.workbook)
        self.ws = self.wf.get_sheet(sheet)

    def set_cell(self,row,col,value):
        #设置单元格数据
        self.ws.write(row,col,value)

    def save_excel(self,filename,format):
        #获取当前时间
        self.time = time.strftime("%Y%m%d%H%M%S", time.localtime())
        #生成文件的文件名及格式
        self.report = filename + '_' +self.time + format
        #保存文件
        self.wf.save(self.report)
'''


